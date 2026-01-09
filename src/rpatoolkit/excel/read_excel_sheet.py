from typing import Any

import polars as pl
from openpyxl import load_workbook
from python_calamine import CalamineWorkbook, SheetVisibleEnum

from rpatoolkit.utils import strip_punctuation


def _get_visible_sheets(path_or_filelike: Any):
    visible = []
    sheets_meta = CalamineWorkbook.from_object(path_or_filelike).sheets_metadata
    for meta in sheets_meta:
        if meta.visible == SheetVisibleEnum.Visible:
            visible.append(meta.name)

    return visible


def read_excel_sheet(
    source: Any,
    *,
    sheet_name: str | None = None,
    header_row: int | None = None,
    visible_rows_only: bool = False,
    lower_column_names: bool = False,
    clean_column_names: bool = False,
    schema_overrides: dict[str, pl.DataType] | None = None,
    schema_override_strict: bool = False,
    raise_if_empty: bool = True,
    drop_empty_rows: bool = False,
    drop_empty_cols: bool = False,
) -> pl.DataFrame:
    """
    Reads a single sheet from an excel file.

    Parameters
    ----------
    source : Any
        Path or File-like object
    sheet_name : str | None, optional
        Name of the sheet to read, by default None. If None, reads the first sheet.
    header_row : int, optional
        Row number to use as header (0-indexed), by default None
    visible_rows_only : bool, optional
        Whether to only read the visible/filtered rows. Uses openpyxl (slower), by default False
    lower_column_names : bool, optional
        Convert column names to lowercase, by default True
    clean_column_names : bool, optional
        Clean column names by stripping punctuation, by default False
    schema_overrides : dict[str, pl.DataType] | None, optional
        Dictionary mapping column names to desired polars data types, by default None
    schema_override_strict : bool, optional
        Whether to raise an error if casting fails, by default False
    raise_if_empty : bool, optional
        Raise an exception if the resulting sheet is empty, by default True
    drop_empty_rows : bool, optional
        Remove empty rows from the sheet, by default False
    drop_empty_cols : bool, optional
        Remove empty columns from the sheet, by default False

    Note:
    -----
    Column names are stripped and converted to lowercase when lower_column_names=True
    """
    if visible_rows_only:
        df = _read_visible_rows_from_sheet(
            source,
            sheet_name=sheet_name,
            header_row=header_row,
        )
    else:
        read_options = {"header_row": header_row} if header_row else None
        df = pl.read_excel(
            source,
            sheet_name=sheet_name,
            read_options=read_options,
            drop_empty_cols=False,
            drop_empty_rows=False,
        )

    if raise_if_empty and df.height == 0:
        raise ValueError(f"No rows found in the sheet: '{sheet_name or 'Default'}'")

    if drop_empty_cols:
        df = df.select([col for col in df.columns if not df[col].is_null().all()])

    if drop_empty_rows:
        df = df.filter(~pl.all_horizontal(pl.all().is_null()))

    if lower_column_names:
        df.columns = [col.strip().lower() for col in df.columns]

    if clean_column_names:
        df.columns = [strip_punctuation(col) for col in df.columns]

    if schema_overrides:
        df = df.cast(
            schema_overrides,
            strict=schema_override_strict,
        )

    return df


def _read_visible_rows_from_sheet(
    source,
    *,
    sheet_name: str | None = None,
    header_row: int | None = None,
):
    """
    Internal helper to read only visible rows using OpenPyXL.
    """
    wb = load_workbook(source, data_only=True)
    if sheet_name:
        try:
            ws = wb[sheet_name]
        except KeyError:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
    else:
        ws = wb[wb.sheetnames[0]]

    visible_rows = []
    for row in ws.iter_rows():
        row_idx = row[0].row
        is_row_hidden = ws.row_dimensions[row_idx].hidden
        if is_row_hidden:
            continue

        values = [cell.value if cell.value != "" else None for cell in row]
        visible_rows.append(values)

    if not visible_rows:
        return pl.DataFrame()

    idx = header_row if header_row else 0
    if idx >= len(visible_rows):
        # Index out of bounds
        raise ValueError(f"header_row index: '{header_row}' out of bounds")

    headers = visible_rows[idx]
    data = visible_rows[idx + 1 :]

    return pl.DataFrame(
        data,
        schema=headers,
        strict=False,
    )
